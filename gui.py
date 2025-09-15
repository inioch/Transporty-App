import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Treeview
from pathlib import Path
import os
from openpyxl import load_workbook

from logic import (
    read_and_filter_csv,
    save_to_excel,
    save_filtered_to_excel,
)


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Transporty")

        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        self.root.geometry(f"{screen_width}x{screen_height}")

        self.file_path = None
        self.table_frame = tk.Frame(self.root)
        self.table_frame.pack(fill="both", expand=True)

        self.sum_parts = 0
        self.setup_gui()

    def setup_gui(self):
        self.packageCount = tk.StringVar()
        self.packageCount.set("Ilość paczek")

        menubar = tk.Menu(self.root)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Wybierz pliki", command=self.select_files)
        file_menu.add_command(label="Wybierz folder", command=self.select_folder)
        file_menu.add_separator()
        file_menu.add_command(label="Exportuj palety i wielopaki", command=self.export_both)
        menubar.add_cascade(label="Plik", menu=file_menu)

        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Pokaż wszystko", command=self.show_table)
        view_menu.add_command(label="Pokaż palety", command=self.show_pallets)
        view_menu.add_command(label="Pokaż wielopaki", command=self.show_multipacks)
        menubar.add_cascade(label="Widok", menu=view_menu)

        self.root.config(menu=menubar)

        self.label_count = tk.Label(self.root, textvariable=self.packageCount, font=("Arial", 12))
        self.label_count.pack(pady=10)

    # === pliki ===
    def delete_file(self, file_name="A_clean.xlsx"):
        plik = Path(file_name)
        if plik.exists():
            plik.unlink()

    def ask_save_location(self, default_name):
        return filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

    def select_files(self):
        os.chdir(os.path.expanduser("~"))
        file_paths = filedialog.askopenfilenames(
            title="Wybierz pliki CSV",
            filetypes=[("CSV Files", "*.csv *.CSV"), ("Wszystkie pliki", "*.*")]
        )

        if not file_paths:
            messagebox.showinfo("Brak plików", "Nie wybrano żadnych plików")
            return

        if messagebox.askyesno("Potwierdzenie", "Czy chcesz przekonwertować wybrane pliki?"):
            self.delete_file()
            self.convert_multiple_csv(file_paths)

    def select_folder(self):
        folder_path = filedialog.askdirectory(title="Wybierz folder z plikami CSV")
        if not folder_path:
            messagebox.showinfo("Brak folderu", "Nie wybrano folderu")
            return

        csv_files = list(Path(folder_path).glob("*.csv"))
        if not csv_files:
            messagebox.showinfo("Brak plików", "Brak plików CSV w wybranym folderze")
            return

        if messagebox.askyesno("Potwierdzenie", f"Znaleziono {len(csv_files)} plików. Czy chcesz je przetworzyć?"):
            self.delete_file()
            self.convert_multiple_csv(csv_files)
            save_to_excel([], [])  # placeholder

    # === logika + GUI ===
    def convert_multiple_csv(self, file_paths):
        selected_headers = ['LP', 'AWB', 'Parts', 'Weight', 'Name', 'Address', 'Town', 'Postcode', 'Number', 'Product']
        selected_columns = [2, 4, 10, 11, 26, 27, 29, 30, 34, 52]

        all_data = []
        self.sum_parts = 0

        for file_path in file_paths:
            data, parts_sum = read_and_filter_csv(file_path, selected_columns, selected_headers)
            if data:
                all_data.extend(data)
                self.sum_parts += parts_sum

        if all_data:
            save_to_excel(selected_headers, all_data)
            messagebox.showinfo("Sukces", "Dane zapisano")
            self.show_table()
            self.package_count()
        else:
            messagebox.showinfo("Brak danych", "Nie znaleziono danych w żadnym z plików")

    # === Tabela ===
    def sort_columns(self, col, reverse):
        def try_float(val):
            try:
                return float(val)
            except ValueError:
                return val

        data = [(try_float(self.tree.set(k, col)), k) for k in self.tree.get_children('')]
        data.sort(reverse=reverse)
        for index, (val, k) in enumerate(data):
            self.tree.move(k, '', index)

        self.tree.heading(col, command=lambda: self.sort_columns(col, not reverse))

    def show_table(self):
        try:
            if not Path("A_clean.xlsx").exists():
                messagebox.showerror("Błąd", "Plik A_clean.xlsx nie istnieje. Najpierw go wygeneruj.")
                return

            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active

            for widget in self.table_frame.winfo_children():
                widget.destroy()

            columns = sheet[1]
            column_names = [cell.value for cell in columns]

            self.tree = Treeview(self.table_frame, columns=column_names, show="headings")
            for col in column_names:
                self.tree.heading(col, text=col, command=lambda c=col: self.sort_columns(c, False))
                self.tree.column(col, width=100)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", "end", values=row)

            self.tree.pack(fill="both", expand=True)

        except Exception as e:
            messagebox.showerror("Błąd", str(e))

    def show_filtered_table(self, headers, data):
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        self.tree = Treeview(self.table_frame, columns=headers, show="headings")
        for col in headers:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_columns(c, False))
            self.tree.column(col, width=100)

        for row in data:
            self.tree.insert("", "end", values=row)

        self.tree.pack(fill="both", expand=True)

    # === licznik paczek ===
    def package_count(self):
        self.packageCount.set(f"Ilość paczek: {self.sum_parts}")

    # === eksporty ===
    def export_pallets(self):
        try:
            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            filtered_rows = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    parts = int(row[2])
                    weight = float(row[3])
                    if parts > 0 and (weight / parts) > 30:
                        filtered_rows.append(row)
                except (ValueError, TypeError):
                    continue

            if not filtered_rows:
                messagebox.showinfo("Brak danych", "Brak palet do eksportu")
                return

            save_path = self.ask_save_location("Palety.xlsx")
            if not save_path:
                return

            save_filtered_to_excel(headers, filtered_rows, Path(save_path))
            messagebox.showinfo("Sukces", f"Plik zapisano jako:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Błąd", str(e))

    def show_pallets(self):
        try:
            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            filtered_rows = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    parts = int(row[2])
                    weight = float(row[3])
                    if parts > 0 and (weight / parts) > 30:
                        filtered_rows.append(row)
                except:
                    continue

            self.show_filtered_table(headers, filtered_rows)

        except Exception as e:
            messagebox.showerror("Błąd", str(e))

    def show_multipacks(self):
        try:
            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            filtered_rows = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    parts = int(row[2])
                    if parts > 10:
                        filtered_rows.append(row)
                except:
                    continue

            self.show_filtered_table(headers, filtered_rows)

        except Exception as e:
            messagebox.showerror("Błąd", str(e))

    def export_both(self):
        try:
            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            pallets, multipacks = [], []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    parts = int(row[2])
                    weight = float(row[3])
                    if parts > 0 and (weight / parts) > 30:
                        pallets.append(row)
                    if parts > 10:
                        multipacks.append(row)
                except (ValueError, TypeError):
                    continue

            if not pallets and not multipacks:
                messagebox.showinfo("Brak danych", "Brak palet i wielopaków do eksportu")
                return

            folder = filedialog.askdirectory(title="Wybierz folder do zapisu")
            if not folder:
                return

            if pallets:
                save_filtered_to_excel(headers, pallets, Path(folder) / "Palety.xlsx")
            if multipacks:
                save_filtered_to_excel(headers, multipacks, Path(folder) / "Wielopaki.xlsx")

            messagebox.showinfo("Sukces", f"Pliki zapisano w:\n{folder}")

        except Exception as e:
            messagebox.showerror("Błąd", str(e))
