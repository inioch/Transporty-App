import csv
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def clean_parts_value(val):
    if isinstance(val, str) and "\\" in val:
        val = val.split("\\")[0]
    try:
        return int(val)
    except ValueError:
        return 0


def read_and_filter_csv(file_path, selected_columns, selected_headers):
    data = []
    with open(file_path, "r", encoding="utf-8") as file:
        reader = csv.reader(file)
        data = [row for row in reader if row and row[0].strip('"') == "A"]

    if not data:
        return None, 0

    filtered_data = []
    sum_parts = 0

    for row in data:
        try:
            selected = []
            for j, idx in enumerate(selected_columns):
                val = row[idx]
                if selected_headers[j] == "Parts":
                    val = clean_parts_value(val)
                    sum_parts += val
                selected.append(val)
            filtered_data.append(selected)
        except (ValueError, IndexError):
            continue

    return filtered_data, sum_parts


def save_to_excel(headers, data, file_name="A_clean.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dane"
    ws.append(headers)

    for row in data:
        ws.append(row)

    for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

        if idx in [1, 2]:
            for cell in col:
                try:
                    cell.value = int(cell.value)
                    cell.number_format = "0"
                except:
                    pass

    wb.save(os.path.join(os.getcwd(), file_name))


def save_filtered_to_excel(headers, data, file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = file_path.stem
    ws.append(headers)

    for row in data:
        ws.append(row)

    for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

    wb.save(file_path)
