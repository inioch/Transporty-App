import csv
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Treeview
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Transporty")

# zawsze pelny ekran
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        self.root.geometry(f"{screen_width}x{screen_height}")   

        self.file_path = None
        self.table_frame = tk.Frame(self.root)
        self.table_frame.pack(fill="both", expand=True)

        self.sum_parts = 0
        self.setup_gui()

    # === GUI Layout ===
    def setup_gui(self):
    # string do pokazywanie ilosci paczek
        self.packageCount = tk.StringVar()
        self.packageCount.set("Ilość paczek")

        menubar = tk.Menu(self.root)

# Menu: Plik

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Wybierz pliki", command=self.select_files)
        file_menu.add_command(label="Wybierz folder", command=self.select_folder)
        file_menu.add_separator()
        file_menu.add_command(label="Exportuj palety i wielopaki", command=self.export_both)
        menubar.add_cascade(label="Plik", menu=file_menu)

# Menu: Widok
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Pokaż wszystko", command=self.show_table)
        view_menu.add_command(label="Pokaż palety", command=self.show_pallets)
        view_menu.add_command(label="Pokaż wielopaki", command=self.show_multipacks)
        menubar.add_cascade(label="Widok", menu=view_menu)

# Menu: dodac o autorze!
        analysis_menu = tk.Menu(menubar, tearoff=0)
        analysis_menu.add_command(label="O autorze", command=self.autor)
        menubar.add_cascade(label="O autorze", menu=analysis_menu)

# Dodanie menu do okna
        self.root.config(menu=menubar)        

# Etykieta z licznikiem paczek
        self.label_count = tk.Label(self.root, textvariable=self.packageCount, font=("Arial", 12))
        self.label_count.pack(pady=10)

    def autor(self):
        messagebox.showinfo("O autorze", "Aplikacja stworzona przez Pawła Kuczaka – kopiowanie i używanie bez wiedzy autora zabronione")
    def convert_multiple_csv(self, file_paths):
        selected_headers = ['LP', 'AWB', 'Parts', 'Weight', 'Name', 'Address', 'Town', 'Postcode', 'Number', 'Product']
        selected_columns = [2, 4, 10, 11, 26, 27, 29, 30, 34, 52]

        all_data = []
        self.sum_parts = 0  # reset

        for file_path in file_paths:
            data = self.read_and_filter_csv(file_path, selected_columns, selected_headers)
            print(file_path, "-> wierszy:", len(data) if data else 0)  # <- debug
            if data:
                all_data.extend(data)

        print("Łącznie wierszy do zapisu:", len(all_data))  # <- debug

        if all_data:
            for idx, row in enumerate(all_data, start=1):
                row[0] = idx
                self.sum_parts += int(row[2])
            self.save_to_excel(selected_headers, all_data)
        else:
            messagebox.showinfo("Brak danych", "Nie znaleziono danych w żadnym z plików")

# pytania gdzie zapisac plik
    def ask_save_location(self, default_name):
        return filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

    # === Obsługa sortowania ===
    def sort_columns(self, col, reverse):
        def try_float(val):
            try:
                return float(val)
            except ValueError:
                return val

        data = [(try_float(self.tree.set(k, col)), k) for k in self.tree.get_children('')]
        data.sort(reverse=reverse)
        for index, (val, k) in enumerate(data):
            self.tree.move(k, '', index)

        self.tree.heading(col, command=lambda: self.sort_columns(col, not reverse))

    def select_folder(self):
        folder_path = filedialog.askdirectory(title="Wybierz folder z plikami CSV")
        if not folder_path:
            messagebox.showinfo("Brak folderu", "Nie wybrano folderu")
            return

        csv_files = list(Path(folder_path).glob("*.csv"))
        if not csv_files:
            messagebox.showinfo("Brak plików", "Brak plików CSV w wybranym folderze")
            return

        if messagebox.askyesno("Potwierdzenie", f"Znaleziono {len(csv_files)} plików. Czy chcesz je przetworzyć?"):
            self.delete_file()
            self.convert_multiple_csv(csv_files)  # <-- tylko to wystarczy

# === Obsługa plików ===
    def delete_file(self, file_name="A_clean.xlsx"):
        plik = Path(file_name)
        if plik.exists():
            plik.unlink()

    def select_files(self):
        os.chdir(os.path.expanduser("~"))
        file_paths = filedialog.askopenfilenames(
        title="Wybierz pliki CSV",
        filetypes=[("CSV Files", "*.csv *.CSV"), ("Wszystkie pliki", "*.*")])

        if not file_paths:
            messagebox.showinfo("Brak plików", "Nie wybrano żadnych plików")
            return

        if messagebox.askyesno("Potwierdzenie", "Czy chcesz przekonwertować wybrane pliki?"):
            self.delete_file()
            self.convert_multiple_csv(file_paths)

# === Logika przetwarzania ===
    def read_and_filter_csv(self, file_path, selected_columns, selected_headers):
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                reader = csv.reader(file)
                data = [row for row in reader if row and row[0].strip('"') == "A"]

            if not data:
                messagebox.showinfo("Brak danych", "Nie znaleziono pozycji 'A' w pliku")
                return None

            filtered_data = []
            self.sum_parts = 0  # reset

            for row in data:
                try:
                    selected = []
                    for j, idx in enumerate(selected_columns):
                        val = row[idx]
                        if selected_headers[j] == "Parts":
                            val =self.clean_parts_value(val)
                            self.sum_parts += val
                        selected.append(val)
                    filtered_data.append(selected)
                except (ValueError, IndexError):
                    continue

            return filtered_data

        except Exception as e:
            messagebox.showerror("Błąd", str(e))
            return None
    def clean_parts_value(self,val):
        if isinstance(val,str) and "\\" in val:
            val = val.split("\\")[0]
        try:
            return int(val)
        except ValueError:
            return 0
    def save_to_excel(self, headers, data, file_name="A_clean.xlsx"):
        file_path = os.path.join(os.getcwd(), file_name)

        if os.path.exists(file_path):
            # Jeśli plik istnieje → wczytaj go
            wb = load_workbook(file_path)
            if "Dane" in wb.sheetnames:
                ws = wb["Dane"]
            else:
                ws = wb.active
        else:
            # Jeśli plik nie istnieje → utwórz nowy
            wb = Workbook()
            ws = wb.active
            ws.title = "Dane"
            ws.append(headers)  # dodaj nagłówki tylko raz, w nowym pliku

        # Dopisz nowe dane
        for row in data:
            ws.append(row)

        # auto szerokość kolumn + formatowanie
        for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

            if idx in [1, 2]:  # LP i AWB jako liczby
                for cell in col:
                    try:
                        cell.value = int(cell.value)
                        cell.number_format = "0"
                    except:
                        pass

        wb.save(file_path)

        messagebox.showinfo("Sukces", f"Plik zapisano jako {file_name}")
        self.show_table()
        self.package_count()


    # === Wyświetlanie tabeli ===
    def show_table(self):
        try:
            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active
            if not Path("A_clean.xlsx").exists():
                messagebox.showerror("Błąd", "Plik A_clean.xlsx nie istnieje. Najpierw go wygeneruj.")
                return


            for widget in self.table_frame.winfo_children():
                widget.destroy()

            columns = sheet[1]
            column_names = [cell.value for cell in columns]

            self.tree = Treeview(self.table_frame, columns=column_names, show="headings")
            for col in column_names:
                self.tree.heading(col, text=col, command=lambda c=col: self.sort_columns(c, False))
                self.tree.column(col, width=100)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", "end", values=row)

            self.tree.pack(fill="both", expand=True)

        except Exception as e:
            messagebox.showerror("Błąd", str(e))

    # === Liczenie paczek ===
    def package_count(self):
        self.packageCount.set(f"Ilość paczek: {self.sum_parts}")
# eksport palet
    def export_pallets(self):
        try:
            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            filtered_rows = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    parts = int(row[2])  # Parts = index 2
                    weight = float(row[3])  # Weight = index 3
                    if parts > 0 and (weight / parts) > 30:
                        filtered_rows.append(row)
                except (ValueError, TypeError):
                    continue

            if not filtered_rows:
                messagebox.showinfo("Brak danych", "Brak palet do eksportu")
                return

            save_path = self.ask_save_location("Palety.xlsx")
            if not save_path:
                return

            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "Palety"
            new_ws.append(headers)
            for row in filtered_rows:
                new_ws.append(row)
            new_wb.save(save_path)
            messagebox.showinfo("Sukces", f"Plik zapisano jako:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Błąd", str(e))

    def show_filtered_table(self, headers, data):
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        self.tree = Treeview(self.table_frame, columns=headers, show="headings")
        for col in headers:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_columns(c, False))
            self.tree.column(col, width=100)

        for row in data:
            self.tree.insert("", "end", values=row)

        self.tree.pack(fill="both", expand=True)

# pokazany w GUI
    def show_pallets(self):
        try:
            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            filtered_rows = []

            # iterujemy po wierszach
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    parts = int(row[2])
                    weight = float(row[3])
                    if parts > 0 and (weight / parts) > 30:  # paleta
                        single_weight = weight / parts
                        # rozbijamy na pojedyncze wiersze
                        for _ in range(parts):
                            new_row = list(row)
                            new_row[0] = 0  # LP zostawimy do ponownego numerowania
                            new_row[2] = 1  # Parts = 1
                            new_row[3] = single_weight  # waga pojedynczej sztuki
                            filtered_rows.append(new_row)
                except:
                    continue

            # przelicz LP od nowa
            for idx, row in enumerate(filtered_rows, start=1):
                row[0] = idx  # LP = idx

            self.show_filtered_table(headers, filtered_rows)

        except Exception as e:
            messagebox.showerror("Błąd", str(e))


    def show_multipacks(self):
        try:
            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            filtered_rows = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    parts = int(row[2])
                    if parts > 10:  # wielopak
                        # zostawiamy w całości
                        new_row = list(row)
                        new_row[0] = 0  # LP = 0, przeliczymy od nowa
                        filtered_rows.append(new_row)
                except:
                    continue

            # przelicz LP od nowa
            for idx, row in enumerate(filtered_rows, start=1):
                row[0] = idx

            self.show_filtered_table(headers, filtered_rows)

        except Exception as e:
            messagebox.showerror("Błąd", str(e))
    def export_both(self):
        try:
            wb = load_workbook("A_clean.xlsx")
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            pallets = []
            multipacks = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    parts = int(row[2])
                    weight = float(row[3])
                    if parts > 0 and (weight / parts) > 30:
                        pallets.append(row)
                    if parts > 10:
                        multipacks.append(row)
                except (ValueError, TypeError):
                    continue

            if not pallets and not multipacks:
                messagebox.showinfo("Brak danych", "Brak palet i wielopaków do eksportu")
                return

            folder = filedialog.askdirectory(title="Wybierz folder do zapisu")
            if not folder:
                return

            if pallets:
                self.save_filtered_to_excel(headers, pallets, Path(folder) / "Palety.xlsx")
            if multipacks:
                self.save_filtered_to_excel(headers, multipacks, Path(folder) / "Wielopaki.xlsx")

            messagebox.showinfo("Sukces", f"Pliki zapisano w:\n{folder}")

        except Exception as e:
            messagebox.showerror("Błąd", str(e))
    def save_filtered_to_excel(self, headers, data, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = file_path.stem
        ws.append(headers)

        for row in data:
            ws.append(row)

        for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

        wb.save(file_path)


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
